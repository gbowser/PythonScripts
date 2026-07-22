#!/usr/bin/env python3
"""Launch an interactive tester from a selected optimisation workbook row."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
import sys
import tkinter as tk
from tkinter import messagebox, ttk
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
for path in (SCRIPT_DIR, PROJECT_ROOT):
    if str(path) not in sys.path:
        sys.path.append(str(path))

from machine_paths import PC_RESEARCH_FOLDERS, remove_foreground_folder  # noqa: E402


WORKBOOK_NAME = "Foreground Masking Optimisation Results.xlsx"
SHEET_TO_TOOL = {
    "MTObjects Spike Gate": "mtobjects_spike_gate",
    "MTObjects Toy Object": "mtobjects",
    "SEP Spike Gate": "sep_spike_gate",
    "SEP Toy Object": "sep",
}
PARAMETER_KEYS = {
    "mtobjects": {
        "move_factor",
        "min_distance",
        "gaussian_fwhm",
        "minarea",
        "dilation_radius",
        "max_area",
        "max_elongation",
    },
    "mtobjects_spike_gate": {
        "move_factor",
        "min_distance",
        "gaussian_fwhm",
        "minarea",
        "dilation_radius",
        "max_area",
        "max_elongation",
    },
    "sep": {
        "detect_thresh",
        "minarea",
        "deblend_nthresh",
        "deblend_cont",
        "back_size",
        "filter_size",
        "dilation_radius",
        "max_area",
        "max_elongation",
    },
    "sep_spike_gate": {
        "detect_thresh",
        "minarea",
        "deblend_nthresh",
        "deblend_cont",
        "back_size",
        "filter_size",
        "dilation_radius",
        "max_area",
        "max_elongation",
    },
}


def default_workbook_path(pc_name: str) -> Path:
    return remove_foreground_folder(pc_name) / "documentation" / WORKBOOK_NAME


def parse_number(value: Any) -> Any:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    try:
        number = float(text)
    except ValueError:
        return value
    if math.isfinite(number) and number.is_integer() and "." not in text and "e" not in text.lower():
        return int(number)
    return number


def row_dicts(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet_name in workbook.sheetnames:
        if sheet_name not in SHEET_TO_TOOL:
            continue
        worksheet = workbook[sheet_name]
        values = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value) if value is not None else "" for value in next(values)]
        except StopIteration:
            continue
        for index, row_values in enumerate(values, start=2):
            row = dict(zip(headers, row_values))
            if not any(value not in (None, "") for value in row.values()):
                continue
            row["_sheet"] = sheet_name
            row["_excel_row"] = index
            row["_tool"] = SHEET_TO_TOOL[sheet_name]
            rows.append(row)
    return rows


def best_params(row: dict[str, Any]) -> dict[str, Any]:
    raw = row.get("best_params_json")
    params: dict[str, Any] = {}
    if raw:
        try:
            loaded = json.loads(str(raw))
            if isinstance(loaded, dict):
                params.update(loaded)
        except json.JSONDecodeError:
            pass

    tool = str(row["_tool"])
    for key in PARAMETER_KEYS[tool]:
        trial_key = f"best_trial_{key}"
        if trial_key in row and row[trial_key] not in (None, ""):
            params[key] = parse_number(row[trial_key])
    optimised_keys = PARAMETER_KEYS[str(row["_tool"])]
    return {
        key: value
        for key, value in params.items()
        if key in optimised_keys and value is not None
    }


def set_if_present(obj: Any, name: str, value: str | None) -> None:
    if value is not None and hasattr(obj, name):
        getattr(obj, name).set(value)


def apply_parameters(app: Any, params: dict[str, Any], row: dict[str, Any]) -> None:
    for key, value in params.items():
        if key not in getattr(app, "vars", {}):
            continue
        app.vars[key].set(app.convert_from_pixels(key, float(value)))
        if hasattr(app, "parameter_changed"):
            app.parameter_changed(key, mark=False)
        if hasattr(app, "format_spinbox_value"):
            app.format_spinbox_value(key)

    set_if_present(app, "detect_on_var", row.get("config_detect_on"))
    set_if_present(app, "spike_gate_detect_on_var", row.get("config_spike_gate_detect_on"))
    if hasattr(app, "mark_needs_calculation"):
        app.mark_needs_calculation()


def launch_interactive(row: dict[str, Any], pc: str, manifest: Path | None, mtobjects_root: Path | None) -> None:
    params = best_params(row)
    tool = str(row["_tool"])

    if tool == "mtobjects_spike_gate":
        import interactive_mtobjects_parameter_tester as module

        app = module.MTObjectsTester(manifest or module.DEFAULT_MANIFEST, pc, mtobjects_root)
    elif tool == "mtobjects":
        import interactive_mtobjects_parameter_tester as module

        app = module.MTObjectsTester(manifest or module.DEFAULT_MANIFEST, pc, mtobjects_root)
    elif tool == "sep_spike_gate":
        import interactive_sep_spike_gate_parameter_tester as module

        app = module.SEPTester(manifest or module.DEFAULT_MANIFEST, pc)
    elif tool == "sep":
        import interactive_sep_parameter_tester as module

        app = module.SEPTester(manifest or module.DEFAULT_MANIFEST, pc)
    else:
        raise ValueError(f"Unsupported tool type: {tool}")

    apply_parameters(app, params, row)
    app.status.set(
        f"Loaded {row['_sheet']} row {row['_excel_row']} from optimisation workbook."
    )
    app.mainloop()


class OptimisationLauncher(tk.Tk):
    def __init__(self, workbook_path: Path, pc: str, manifest: Path | None, mtobjects_root: Path | None) -> None:
        super().__init__()
        self.title("Launch Interactive Tool from Optimisation Results")
        self.geometry("1180x520")
        self.minsize(920, 420)
        self.workbook_path = workbook_path
        self.pc = pc
        self.manifest = manifest
        self.mtobjects_root = mtobjects_root
        self.rows = row_dicts(workbook_path)
        self.filtered_rows = list(self.rows)

        self.search_var = tk.StringVar()
        self.summary_var = tk.StringVar(value=f"{len(self.rows)} optimisation runs loaded from {workbook_path}")
        self._build()
        self._populate()

    def _build(self) -> None:
        outer = ttk.Frame(self, padding=10)
        outer.pack(fill=tk.BOTH, expand=True)

        top = ttk.Frame(outer)
        top.pack(fill=tk.X, pady=(0, 8))
        ttk.Label(top, text="Filter").pack(side=tk.LEFT)
        search = ttk.Entry(top, textvariable=self.search_var)
        search.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(6, 8))
        search.bind("<KeyRelease>", lambda _event: self._filter())
        ttk.Button(top, text="Refresh", command=self._refresh).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(top, text="Launch Selected", command=self._launch_selected).pack(side=tk.LEFT)

        columns = (
            "sheet",
            "row",
            "run_datetime",
            "total_trials",
            "successful_trials",
            "best_objective",
            "best_score",
            "params",
        )
        self.tree = ttk.Treeview(outer, columns=columns, show="headings", selectmode="browse")
        headings = {
            "sheet": "Optimiser",
            "row": "Excel row",
            "run_datetime": "Run date/time",
            "total_trials": "Trials",
            "successful_trials": "OK",
            "best_objective": "Best objective",
            "best_score": "Best score",
            "params": "Best parameters",
        }
        widths = {
            "sheet": 170,
            "row": 70,
            "run_datetime": 150,
            "total_trials": 70,
            "successful_trials": 70,
            "best_objective": 115,
            "best_score": 100,
            "params": 390,
        }
        for column in columns:
            self.tree.heading(column, text=headings[column])
            self.tree.column(column, width=widths[column], minwidth=50, stretch=(column == "params"))
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree.bind("<Double-1>", lambda _event: self._launch_selected())

        y_scroll = ttk.Scrollbar(outer, orient=tk.VERTICAL, command=self.tree.yview)
        y_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.configure(yscrollcommand=y_scroll.set)

        ttk.Label(self, textvariable=self.summary_var, anchor=tk.W).pack(fill=tk.X, padx=10, pady=(0, 8))

    def _refresh(self) -> None:
        self.rows = row_dicts(self.workbook_path)
        self._filter()

    def _filter(self) -> None:
        needle = self.search_var.get().strip().lower()
        if not needle:
            self.filtered_rows = list(self.rows)
        else:
            self.filtered_rows = [
                row for row in self.rows
                if needle in " ".join(str(value) for value in row.values()).lower()
            ]
        self._populate()

    def _populate(self) -> None:
        self.tree.delete(*self.tree.get_children())
        for index, row in enumerate(self.filtered_rows):
            params = best_params(row)
            params_text = ", ".join(f"{key}={value:.4g}" if isinstance(value, float) else f"{key}={value}" for key, value in params.items())
            self.tree.insert(
                "",
                tk.END,
                iid=str(index),
                values=(
                    row.get("_sheet", ""),
                    row.get("_excel_row", ""),
                    row.get("run_datetime", ""),
                    row.get("total_trials", ""),
                    row.get("successful_trials", ""),
                    row.get("best_objective", ""),
                    row.get("best_score", ""),
                    params_text,
                ),
            )
        self.summary_var.set(f"{len(self.filtered_rows)} of {len(self.rows)} optimisation runs shown from {self.workbook_path}")

    def _launch_selected(self) -> None:
        selection = self.tree.selection()
        if not selection:
            messagebox.showinfo("No row selected", "Select an optimisation run first.")
            return
        row = self.filtered_rows[int(selection[0])]
        self.destroy()
        launch_interactive(row, self.pc, self.manifest, self.mtobjects_root)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--pc", choices=sorted(PC_RESEARCH_FOLDERS), default="Desktop")
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--manifest", type=Path, default=None)
    parser.add_argument("--mtobjects-root", type=Path, default=None)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook or default_workbook_path(args.pc)
    if not workbook_path.is_file():
        raise FileNotFoundError(f"Optimisation results workbook not found: {workbook_path}")
    app = OptimisationLauncher(workbook_path, args.pc, args.manifest, args.mtobjects_root)
    app.mainloop()


if __name__ == "__main__":
    main()
