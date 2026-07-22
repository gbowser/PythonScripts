#!/usr/bin/env python3
"""Maintain the shared optimisation-results workbook.

Each optimiser writes its normal CSV/JSON artefacts first. This helper then
adds one summary row per optimiser run to a shared XLSX workbook with one
worksheet per algorithm/method pair. The workbook is deliberately
denormalised: every run row carries run metadata, best parameters, aggregate
results, source files, and run date/time so it remains useful even if viewed
away from the output folders.
"""

from __future__ import annotations

import csv
from datetime import datetime
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


WORKBOOK_NAME = "Foreground Masking Optimisation Results.xlsx"
SHEET_NAMES = {
    ("SEP", "Spike Gate"): "SEP Spike Gate",
    ("SEP", "Toy Object"): "SEP Toy Object",
    ("MTObjects", "Spike Gate"): "MTObjects Spike Gate",
    ("MTObjects", "Toy Object"): "MTObjects Toy Object",
}
BASE_COLUMNS = [
    "algorithm",
    "method",
    "run_datetime",
    "run_dir",
    "total_trials",
    "successful_trials",
    "failed_trials",
    "source_summary_csv",
    "source_details_csv",
    "source_best_json",
    "source_config_json",
    "source_cases_or_toys_csv",
    "best_objective",
    "best_score",
    "best_params_json",
]


def default_workbook_path(pc_name: str | None = None) -> Path:
    from machine_paths import remove_foreground_folder

    return remove_foreground_folder(pc_name or "Desktop") / "documentation" / WORKBOOK_NAME


def _timestamp_from_run_dir(run_dir: Path) -> str:
    name = run_dir.name
    try:
        return datetime.strptime(name, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        for part in reversed(run_dir.parts):
            try:
                return datetime.strptime(part, "%Y%m%d_%H%M%S").strftime("%Y-%m-%d %H:%M:%S")
            except ValueError:
                continue
    return datetime.fromtimestamp(run_dir.stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")


def _read_json(path: Path | None) -> dict[str, Any]:
    if path is None or not path.is_file():
        return {}
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def _read_csv(path: Path) -> list[dict[str, Any]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, sort_keys=True)
    return str(value)


def _coerce_cell_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value)
    if text == "":
        return ""
    try:
        if text.lower() in {"nan", "inf", "-inf"}:
            return text
        number = float(text)
        if number.is_integer() and "." not in text and "e" not in text.lower():
            return int(number)
        return number
    except ValueError:
        return text


def _is_success_row(row: dict[str, Any]) -> bool:
    return str(row.get("status", "ok")).strip().lower() == "ok"


def _best_summary_row(summary_rows: list[dict[str, Any]]) -> dict[str, Any]:
    def objective(row: dict[str, Any]) -> float:
        try:
            return float(row.get("objective", "inf"))
        except (TypeError, ValueError):
            return float("inf")

    return min(summary_rows, key=objective)


def _study_files(run_dir: Path, prefix: str) -> dict[str, Path | None]:
    cases_or_toys = list(run_dir.glob(f"{prefix}*_cases.csv")) + list(run_dir.glob(f"{prefix}*_toys.csv"))
    return {
        "summary": next(iter(run_dir.glob(f"{prefix}*_summary.csv")), None),
        "details": next(iter(run_dir.glob(f"{prefix}*_details.csv")), None),
        "best": next(iter(run_dir.glob(f"{prefix}*_best.json")), None),
        "config": next(iter(run_dir.glob(f"{prefix}*_config.json")), None),
        "cases_or_toys": cases_or_toys[0] if cases_or_toys else None,
    }


def _load_or_create(path: Path) -> Workbook:
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.is_file():
        return load_workbook(path)
    workbook = Workbook()
    default = workbook.active
    workbook.remove(default)
    return workbook


def _sheet(workbook: Workbook, algorithm: str, method: str):
    name = SHEET_NAMES[(algorithm, method)]
    if name not in workbook.sheetnames:
        worksheet = workbook.create_sheet(name)
        worksheet.sheet_view.showGridLines = False
        return worksheet
    return workbook[name]


def _existing_run_dirs(worksheet) -> set[str]:
    if worksheet.max_row < 2:
        return set()
    headers = [cell.value for cell in worksheet[1]]
    try:
        run_dir_col = headers.index("run_dir") + 1
    except ValueError:
        return set()
    return {str(worksheet.cell(row=row, column=run_dir_col).value) for row in range(2, worksheet.max_row + 1)}


def _write_rows(worksheet, rows: list[dict[str, Any]]) -> None:
    existing_headers = [cell.value for cell in worksheet[1]] if worksheet.max_row else []
    existing_headers = [header for header in existing_headers if header]
    all_headers = list(existing_headers)
    for header in BASE_COLUMNS:
        if header not in all_headers:
            all_headers.append(header)
    for row in rows:
        for header in row:
            if header not in all_headers:
                all_headers.append(header)

    worksheet.delete_rows(1, worksheet.max_row or 1)
    worksheet.append(all_headers)
    for row in rows:
        worksheet.append([_coerce_cell_value(row.get(header, "")) for header in all_headers])
    _format_sheet(worksheet)


def _format_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="E8EEF5")
    header_font = Font(bold=True, color="0B2545")
    thin = Side(style="thin", color="D9DEE8")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = Border(bottom=thin)
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        letter = get_column_letter(column_cells[0].column)
        header = str(column_cells[0].value or "")
        max_len = max(len(str(cell.value or "")) for cell in column_cells[: min(len(column_cells), 200)])
        width = min(max(max_len + 2, 10), 45)
        if header in {"run_dir", "source_summary_csv", "source_details_csv", "source_best_json", "source_config_json", "best_params_json"}:
            width = min(max(width, 24), 60)
        worksheet.column_dimensions[letter].width = width
        for cell in column_cells[1:]:
            cell.alignment = Alignment(wrap_text=False, vertical="top")


def append_run_to_workbook(
    *,
    algorithm: str,
    method: str,
    run_dir: Path,
    prefix: str,
    workbook_path: Path | None = None,
) -> Path:
    run_dir = Path(run_dir)
    workbook_path = workbook_path or default_workbook_path()
    files = _study_files(run_dir, prefix)
    summary_path = files["summary"]
    if summary_path is None or not summary_path.is_file():
        raise FileNotFoundError(f"No summary CSV found for {algorithm} {method}: {run_dir}")

    all_summary_rows = _read_csv(summary_path)
    if len(all_summary_rows) < 5:
        return workbook_path
    summary_rows = [row for row in all_summary_rows if _is_success_row(row)]
    if not summary_rows:
        return workbook_path

    best_summary = _best_summary_row(summary_rows)
    best = _read_json(files["best"])
    config = _read_json(files["config"])
    best_params = best.get("params", {})
    run_datetime = _timestamp_from_run_dir(run_dir)
    metadata = {
        "algorithm": algorithm,
        "method": method,
        "run_datetime": run_datetime,
        "run_dir": str(run_dir),
        "total_trials": len(all_summary_rows),
        "successful_trials": len(summary_rows),
        "failed_trials": len(all_summary_rows) - len(summary_rows),
        "source_summary_csv": str(summary_path),
        "source_details_csv": str(files["details"] or ""),
        "source_best_json": str(files["best"] or ""),
        "source_config_json": str(files["config"] or ""),
        "source_cases_or_toys_csv": str(files["cases_or_toys"] or ""),
        "best_objective": best.get("objective", ""),
        "best_score": best.get("score", ""),
        "best_params_json": _stringify(best_params),
    }
    for key, value in sorted(config.items()):
        metadata[f"config_{key}"] = _stringify(value)
    for key, value in best_summary.items():
        metadata[f"best_trial_{key}"] = value
    rows = [metadata]

    workbook = _load_or_create(workbook_path)
    worksheet = _sheet(workbook, algorithm, method)
    existing = _existing_run_dirs(worksheet)
    if str(run_dir) in existing:
        # Rebuild without the old copy of this run so reruns update cleanly.
        headers = [cell.value for cell in worksheet[1]] if worksheet.max_row else []
        old_rows = []
        for values in worksheet.iter_rows(min_row=2, values_only=True):
            row = dict(zip(headers, values))
            if str(row.get("run_dir", "")) != str(run_dir):
                old_rows.append(row)
        rows = old_rows + rows
    else:
        headers = [cell.value for cell in worksheet[1]] if worksheet.max_row else []
        old_rows = [dict(zip(headers, values)) for values in worksheet.iter_rows(min_row=2, values_only=True)] if headers else []
        rows = old_rows + rows

    _write_rows(worksheet, rows)
    workbook.save(workbook_path)
    return workbook_path


def backfill_workbook(studies: list[dict[str, str | Path]], workbook_path: Path | None = None) -> Path:
    workbook_path = workbook_path or default_workbook_path()
    for study in studies:
        append_run_to_workbook(
            algorithm=str(study["algorithm"]),
            method=str(study["method"]),
            run_dir=Path(study["run_dir"]),
            prefix=str(study["prefix"]),
            workbook_path=workbook_path,
        )
    return workbook_path
