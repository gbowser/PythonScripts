#!/usr/bin/env python3
"""Build a parameter-stability workbook from optimiser best JSON files."""

from __future__ import annotations

import argparse
import json
import math
from pathlib import Path
from statistics import mean, pstdev
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


METHODS = {
    "SEP Spike Gate": {
        "prefix": "sep_spike_gate",
        "best": "sep_spike_optimisation_best.json",
        "order": [
            "best_objective",
            "detect_thresh",
            "minarea",
            "deblend_nthresh",
            "deblend_cont",
            "back_size",
            "filter_size",
            "dilation_radius",
            "max_area",
            "max_elongation",
        ],
    },
    "SEP Toy Object": {
        "prefix": "sep_toy_objects",
        "best": "sep_toy_object_optimisation_best.json",
        "order": [
            "best_objective",
            "best_score",
            "detect_thresh",
            "minarea",
            "deblend_nthresh",
            "deblend_cont",
            "back_size",
            "filter_size",
            "dilation_radius",
            "max_area",
            "max_elongation",
        ],
    },
    "MTObjects Spike Gate": {
        "prefix": "spike_gate_MTObjects",
        "best": "mtobjects_spike_optimisation_best.json",
        "order": [
            "best_objective",
            "move_factor",
            "min_distance",
            "gaussian_fwhm",
            "bg_variance",
            "minarea",
            "dilation_radius",
            "max_area",
            "max_elongation",
        ],
    },
    "MTObjects Toy Object": {
        "prefix": "toy_objects_MTObjects",
        "best": "mtobjects_parameter_optimisation_best.json",
        "order": [
            "best_objective",
            "best_score",
            "move_factor",
            "min_distance",
            "gaussian_fwhm",
            "bg_variance",
            "minarea",
            "dilation_radius",
            "max_area",
            "max_elongation",
        ],
    },
}


def numeric(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def load_run(best_path: Path) -> dict[str, Any]:
    with best_path.open("r", encoding="utf-8") as handle:
        best = json.load(handle)
    params = best.get("params") if isinstance(best.get("params"), dict) else {}
    row = dict(params)
    if "objective" in best:
        row["best_objective"] = best["objective"]
    if "score" in best:
        row["best_score"] = best["score"]
    row["_run_dir"] = str(best_path.parent)
    return row


def run_columns(run_root: Path, prefix: str, best_name: str) -> list[tuple[str, Path]]:
    runs = []
    for seed_dir in sorted(run_root.glob("seed_*")):
        best_paths = sorted((seed_dir / prefix).glob(f"**/{best_name}"))
        if best_paths:
            runs.append((seed_dir.name.removeprefix("seed_"), best_paths[-1]))
    return runs


def ordered_parameters(order: list[str], runs: list[dict[str, Any]]) -> list[str]:
    seen = set()
    params = []
    for name in order:
        if any(name in run for run in runs):
            params.append(name)
            seen.add(name)
    extras = sorted({key for run in runs for key in run if not key.startswith("_")} - seen)
    return params + extras


def autosize(ws) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        width = min(max(max(len(str(cell.value or "")) for cell in column) + 2, 12), 42)
        ws.column_dimensions[letter].width = width


def build_workbook(run_root: Path, output_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="E8EEF5")
    header_font = Font(bold=True)

    for sheet_name, spec in METHODS.items():
        columns = run_columns(run_root, spec["prefix"], spec["best"])
        if not columns:
            raise FileNotFoundError(
                f"No {spec['best']} files found below {run_root}/seed_*/{spec['prefix']}"
            )
        loaded = [(seed, load_run(path)) for seed, path in columns]
        ws = wb.create_sheet(sheet_name)
        headers = ["parameter", *[seed for seed, _run in loaded], "mean", "stdevp"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for parameter in ordered_parameters(spec["order"], [run for _seed, run in loaded]):
            values = [run.get(parameter, "") for _seed, run in loaded]
            numeric_values = [number for number in (numeric(value) for value in values) if number is not None]
            row_mean = mean(numeric_values) if numeric_values else ""
            row_stdevp = pstdev(numeric_values) if len(numeric_values) >= 2 else (0.0 if len(numeric_values) == 1 else "")
            ws.append([parameter, *values, row_mean, row_stdevp])

        ws.append([])
        ws.append(["run_dir", *[run.get("_run_dir", "") for _seed, run in loaded]])
        ws.freeze_panes = "B2"
        autosize(ws)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--run-root", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    build_workbook(args.run_root, args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
