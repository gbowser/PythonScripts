from __future__ import annotations

import getpass
import re
from pathlib import Path


LOGIN_URL = "https://client.evelyn.com/en/portfolios/list"
WORKBOOK_PATH = Path(r"D:\Dropbox\My Documents\Home Finances\xGPB Portfolio 2026.xlsm")
SHEET_NAME = "Pensions"
VALUE_COLUMN = "I"
DATE_COLUMN = "H"
DEFAULT_USERNAME = "gordonbowser"


def require_dependencies():
    try:
        from openpyxl import load_workbook
        from selenium import webdriver
        from selenium.common.exceptions import TimeoutException
        from selenium.webdriver.common.by import By
        from selenium.webdriver.chrome.options import Options as ChromeOptions
        from selenium.webdriver.chrome.service import Service as ChromeService
        from selenium.webdriver.common.keys import Keys
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.support.ui import WebDriverWait
    except ImportError as exc:
        raise SystemExit(
            "This script needs 'selenium' and 'openpyxl' installed in the current "
            "Python environment. Example: pip install selenium openpyxl"
        ) from exc

    return {
        "load_workbook": load_workbook,
        "webdriver": webdriver,
        "TimeoutException": TimeoutException,
        "By": By,
        "ChromeOptions": ChromeOptions,
        "ChromeService": ChromeService,
        "Keys": Keys,
        "EC": EC,
        "WebDriverWait": WebDriverWait,
    }


def build_driver(deps):
    options = deps["ChromeOptions"]()
    driver = deps["webdriver"].Chrome(options=options)
    driver.set_window_size(1200, 900)
    driver.set_window_position(80, 60)
    return driver


def first_present(wait, locators):
    for by, value in locators:
        try:
            return wait.until(wait._driver.find_element(by, value) if False else None)
        except Exception:
            pass

    by, value = locators[-1]
    return wait.until(lambda driver: driver.find_element(by, value))


def wait_for_any(driver, wait, locators):
    last_error = None
    for by, value in locators:
        try:
            return wait.until(lambda d: d.find_element(by, value))
        except Exception as exc:
            last_error = exc
    raise last_error if last_error else RuntimeError("No matching element found.")


def fill_if_found(wait, locators, value):
    element = wait_for_any(wait._driver, wait, locators)
    element.clear()
    element.send_keys(value)


def click_if_found(wait, locators):
    ec = require_dependencies()["EC"]
    last_error = None
    for by, value in locators:
        try:
            element = wait.until(ec.element_to_be_clickable((by, value)))
            element.click()
            return
        except Exception as exc:
            last_error = exc
    raise last_error if last_error else RuntimeError("No clickable element found.")


def clean_currency(value_text: str) -> float:
    cleaned = re.sub(r"[^\d.\-]", "", value_text)
    return float(cleaned)


def extract_text_near_label(driver, wait, by, locators):
    element = wait_for_any(driver, wait, locators)
    return element.text.strip()


def login_and_fetch_values():
    deps = require_dependencies()
    driver = build_driver(deps)
    wait = deps["WebDriverWait"](driver, 30)
    by = deps["By"]
    keys = deps["Keys"]

    username = input(f"Username [{DEFAULT_USERNAME}]: ").strip() or DEFAULT_USERNAME
    password = getpass.getpass("Password: ")
    verification_code = input(
        "Verification Code: "
    ).strip()

    try:
        driver.get(LOGIN_URL)

        fill_if_found(
            wait,
            [
                (by.NAME, "username"),
                (by.ID, "username"),
                (by.CSS_SELECTOR, "input[type='email']"),
                (by.CSS_SELECTOR, "input[name*='user']"),
            ],
            username,
        )
        password_element = wait_for_any(
            driver,
            wait,
            [
                (by.NAME, "password"),
                (by.ID, "password"),
                (by.CSS_SELECTOR, "input[type='password']"),
            ],
        )
        password_element.clear()
        password_element.send_keys(password)
        password_element.send_keys(keys.ENTER)

        verification_code_locators = [
            (by.NAME, "verificationCode"),
            (by.NAME, "code"),
            (by.ID, "verificationCode"),
            (by.CSS_SELECTOR, "input[inputmode='numeric']"),
            (by.CSS_SELECTOR, "input[name*='code']"),
        ]
        wait_for_any(driver, wait, verification_code_locators)
        fill_if_found(
            wait,
            verification_code_locators,
            verification_code,
        )
        click_if_found(
            wait,
            [
                (by.CSS_SELECTOR, "button[type='submit']"),
                (by.XPATH, "//button[contains(., 'Submit')]"),
                (by.XPATH, "//button[contains(., 'Verify')]"),
            ],
        )

        total_value_text = extract_text_near_label(
            driver,
            wait,
            by,
            [
                (
                    by.XPATH,
                    "//*[contains(normalize-space(), 'Total Value')]/following::*[1]",
                ),
                (
                    by.XPATH,
                    "//*[contains(normalize-space(), 'Total Value')]/ancestor::*[1]"
                    "/following-sibling::*[1]",
                ),
            ],
        )

        last_updated_text = extract_text_near_label(
            driver,
            wait,
            by,
            [
                (
                    by.XPATH,
                    "//*[contains(normalize-space(), 'Last updated')]/following::*[1]",
                ),
                (
                    by.XPATH,
                    "//*[contains(normalize-space(), 'Last updated')]/ancestor::*[1]"
                    "/following-sibling::*[1]",
                ),
            ],
        )

        return clean_currency(total_value_text), last_updated_text
    finally:
        driver.quit()


def find_next_empty_row(sheet, column_index: int) -> int:
    for row in range(sheet.max_row, 0, -1):
        if sheet.cell(row=row, column=column_index).value not in (None, ""):
            return row + 1
    return 1


def update_workbook(total_value: float, last_updated: str):
    deps = require_dependencies()
    load_workbook = deps["load_workbook"]

    workbook = load_workbook(WORKBOOK_PATH, keep_vba=True)
    sheet = workbook[SHEET_NAME]

    value_column_index = sheet[VALUE_COLUMN + "1"].column
    date_column_index = sheet[DATE_COLUMN + "1"].column
    target_row = find_next_empty_row(sheet, value_column_index)

    sheet.cell(row=target_row, column=date_column_index).value = last_updated
    sheet.cell(row=target_row, column=value_column_index).value = total_value

    workbook.save(WORKBOOK_PATH)
    return target_row


def main():
    total_value, last_updated = login_and_fetch_values()
    target_row = update_workbook(total_value, last_updated)
    print(f"Inserted Total Value {total_value} into {SHEET_NAME}!{VALUE_COLUMN}{target_row}")
    print(f"Inserted Last updated {last_updated} into {SHEET_NAME}!{DATE_COLUMN}{target_row}")
    print(f"Last updated date found: {last_updated}")


if __name__ == "__main__":
    main()
