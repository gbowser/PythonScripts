#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Merge shoulder-recognition classifications into the PE/VPD classification workbook.

The input CSV is produced by Real Galaxy Shoulder Quantification v0.69.py as
shoulder_classifications.csv. The workbook is updated in place after a backup is
created next to the original file.
"""

from __future__ import annotations

import argparse
import csv
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
BARPROFILE_DATA_DIR = PROJECT_ROOT / "Erwin_barprofiles_paper_GB_working_copy" / "data"
SCRAMBLED_MAP = BARPROFILE_DATA_DIR / "scrambled_map.txt"
PE_CLASSIFICATIONS = BARPROFILE_DATA_DIR / "classifications_pe.txt"
VPD_CLASSIFICATIONS = BARPROFILE_DATA_DIR / "classifications_vd_revised.txt"

DEFAULT_WORKBOOK = Path(
    r"C:\Users\gordo\Dropbox\Public Documents\UCLAN\MSc Research"
    r"\Shoulder_Recognition_Erwin\PE_VPD_galaxy_classifications_with_definitions.xlsx"
)
PC_RESEARCH_FOLDERS = {
    "Laptop": Path(r"C:\Users\gordo\Dropbox\Public Documents\UCLAN\MSc Research"),
    "Desktop": Path(r"D:\Dropbox\Public Documents\UCLAN\MSc Research"),
}

OUTPUT_COLUMNS = [
    "PE profile class",
    "PE profile label",
    "VPD profile class",
    "VPD profile label",
    "sra_classification",
    "sra_classification_detail",
    "left_shoulder_found",
    "right_shoulder_found",
    "failed_extrema",
    "d_extrema",
    "d2_extrema",
    "roc_minima",
]

GALAXY_COLUMN_CANDIDATES = {
    "galaxy",
    "galaxy_name",
    "galaxyname",
    "name",
    "ngc",
    "object",
    "objectname",
}

PROFILE_CLASS_LABELS = {
    "BP": "Peak+Sh",
    "EXP": "Exp",
    "EXP(N)": "Exp",
    "FT": "Flat-top (FT)",
    "FT(N)": "Flat-top (FT)",
    "FLAT-TOP": "Flat-top (FT)",
    "FLAT-TOP(N)": "Flat-top (FT)",
    "2S": "Two-slope (2S)",
    "2S(N)": "Two-slope (2S)",
    "TWO-SLOPE": "Two-slope (2S)",
    "TWO-SLOPE(N)": "Two-slope (2S)",
}

PROFILE_DEFINITIONS = [
    ("Peak+Sh", "BP", 'Classic EE85 "flat" bar profile associated with B/P bulges; labelled P+Sh in the paper.'),
    ("Exp", "Exp, Exp(N)", "Single-exponential profile; (N) keeps the nuclear-excess qualifier in the raw code."),
    ("Flat-top (FT)", "FT, FT(N), Flat-top", "Roughly constant central surface brightness followed by an exponential falloff."),
    ("Two-slope (2S)", "2S, 2S(N), Two-slope", "Outer steep exponential and inner shallow exponential."),
]


def normalise_name(value: object) -> str:
    return str(value or "").strip().upper().replace(" ", "")


def normalise_profile_code(value: object) -> str:
    code = str(value or "").strip()
    code = code.rstrip("?").replace("(?)", "")
    return code


def profile_label(value: object) -> str:
    code = normalise_profile_code(value)
    return PROFILE_CLASS_LABELS.get(code.upper(), code)


def read_descramble_map(path: Path) -> dict[int, str]:
    mapping = {}
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if not line.strip() or line.startswith("#"):
                continue
            parts = line.split()
            mapping[int(parts[0])] = parts[2]
    return mapping


def read_profile_classifications(path: Path, descramble_map: dict[int, str]) -> dict[str, str]:
    classifications = {}
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            if not line.strip() or line.startswith("#"):
                continue
            parts = line.split()
            if len(parts) < 2 or parts[1] == "?":
                continue
            galaxy = descramble_map[int(parts[0])]
            classifications[normalise_name(galaxy)] = normalise_profile_code(parts[1])
    return classifications


def default_csv_path(pc_name: str) -> Path:
    return PC_RESEARCH_FOLDERS[pc_name] / "Shoulder_Recognition_Erwin" / "shoulder_classifications.csv"


def default_workbook_path(pc_name: str) -> Path:
    return (
        PC_RESEARCH_FOLDERS[pc_name]
        / "Shoulder_Recognition_Erwin"
        / "PE_VPD_galaxy_classifications_with_definitions.xlsx"
    )


def read_classifications(csv_path: Path) -> dict[str, dict[str, str]]:
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.DictReader(handle))

    if not rows:
        raise ValueError(f"No classification rows found in {csv_path}")
    if "galaxy" not in rows[0]:
        raise ValueError(f"{csv_path} must contain a 'galaxy' column")

    classifications = {}
    for row in rows:
        key = normalise_name(row.get("galaxy"))
        if key:
            classifications[key] = row
    return classifications


def find_header_row_and_galaxy_column(sheet):
    for row_index in range(1, min(sheet.max_row, 20) + 1):
        headers = [
            normalise_name(sheet.cell(row=row_index, column=col_index).value).lower()
            for col_index in range(1, sheet.max_column + 1)
        ]
        for col_index, header in enumerate(headers, start=1):
            if header in GALAXY_COLUMN_CANDIDATES:
                return row_index, col_index
    raise ValueError(
        f"Could not find a galaxy/name column in worksheet '{sheet.title}'."
    )


def ensure_output_columns(sheet, header_row: int) -> dict[str, int]:
    existing = {}
    for col_index in range(1, sheet.max_column + 1):
        value = sheet.cell(row=header_row, column=col_index).value
        if value:
            existing[str(value).strip()] = col_index

    for column_name in OUTPUT_COLUMNS:
        if column_name not in existing:
            new_column = sheet.max_column + 1
            sheet.cell(row=header_row, column=new_column).value = column_name
            existing[column_name] = new_column

    return existing


def write_profile_definitions(workbook) -> None:
    sheet = workbook["Definitions"] if "Definitions" in workbook.sheetnames else workbook.create_sheet("Definitions")

    for row_index in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=1).value == "Barprofiles Figure Legend Classes":
            return

    start_row = sheet.max_row + 2 if sheet.max_row else 1
    sheet.cell(row=start_row, column=1).value = "Barprofiles Figure Legend Classes"
    sheet.cell(row=start_row + 2, column=1).value = "Display label"
    sheet.cell(row=start_row + 2, column=2).value = "Source code(s)"
    sheet.cell(row=start_row + 2, column=3).value = "Meaning"

    for offset, (display_label, source_codes, meaning) in enumerate(PROFILE_DEFINITIONS, start=3):
        row = start_row + offset
        sheet.cell(row=row, column=1).value = display_label
        sheet.cell(row=row, column=2).value = source_codes
        sheet.cell(row=row, column=3).value = meaning


def merge_into_workbook(workbook_path: Path, csv_path: Path, sheet_name: str | None) -> int:
    classifications = read_classifications(csv_path)
    descramble_map = read_descramble_map(SCRAMBLED_MAP)
    pe_classifications = read_profile_classifications(PE_CLASSIFICATIONS, descramble_map)
    vpd_classifications = read_profile_classifications(VPD_CLASSIFICATIONS, descramble_map)

    workbook = load_workbook(workbook_path)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    header_row, galaxy_column = find_header_row_and_galaxy_column(sheet)
    output_columns = ensure_output_columns(sheet, header_row)

    matched = 0
    for row_index in range(header_row + 1, sheet.max_row + 1):
        galaxy_key = normalise_name(sheet.cell(row=row_index, column=galaxy_column).value)
        if not galaxy_key or galaxy_key not in classifications:
            continue

        matched += 1
        result_row = classifications[galaxy_key]
        pe_class = pe_classifications.get(galaxy_key, "")
        vpd_class = vpd_classifications.get(galaxy_key, "")
        result_row = {
            **result_row,
            "PE profile class": pe_class,
            "PE profile label": profile_label(pe_class),
            "VPD profile class": vpd_class,
            "VPD profile label": profile_label(vpd_class),
        }
        for column_name in OUTPUT_COLUMNS:
            sheet.cell(
                row=row_index,
                column=output_columns[column_name],
            ).value = result_row.get(column_name)

    backup_path = workbook_path.with_name(
        f"{workbook_path.stem}.backup_{datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup_path)
    write_profile_definitions(workbook)
    workbook.save(workbook_path)
    return matched


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Merge SRA shoulder classifications into the PE/VPD workbook."
    )
    parser.add_argument(
        "--pc",
        choices=sorted(PC_RESEARCH_FOLDERS),
        default="Laptop",
        help="Select which Dropbox research-folder location to use.",
    )
    parser.add_argument(
        "--csv",
        type=Path,
        default=None,
        help="Path to shoulder_classifications.csv.",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help="Path to PE_VPD_galaxy_classifications_with_definitions.xlsx.",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Worksheet name to update. Defaults to the active sheet.",
    )
    args = parser.parse_args()
    if args.csv is None:
        args.csv = default_csv_path(args.pc)
    if args.workbook is None:
        args.workbook = default_workbook_path(args.pc)
    return args


def main() -> int:
    args = parse_args()
    if not args.csv.exists():
        raise FileNotFoundError(f"Could not find classification CSV: {args.csv}")
    if not args.workbook.exists():
        raise FileNotFoundError(f"Could not find workbook: {args.workbook}")

    matched = merge_into_workbook(args.workbook, args.csv, args.sheet)
    print(f"Updated {matched} workbook rows from {args.csv}")
    print(f"Workbook saved: {args.workbook}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
