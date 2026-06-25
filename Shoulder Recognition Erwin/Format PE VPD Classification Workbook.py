#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Reorder and style the PE/VPD/SRA classification workbook."""

from __future__ import annotations

import argparse
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


PC_RESEARCH_FOLDERS = {
    "Laptop": Path(r"C:\Users\gordo\Dropbox\Public Documents\UCLAN\MSc Research"),
    "Desktop": Path(r"D:\Dropbox\Public Documents\UCLAN\MSc Research"),
}

WORKBOOK_NAME = "PE_VPD_galaxy_classifications_with_definitions.xlsx"
VISUAL_REVIEW_CLASS_OPTIONS = [
    "Peak+Sh",
    "Exp",
    "Flat-top (FT)",
    "Two-slope (2S)",
    "Unclear",
]

PREFERRED_COLUMNS = [
    "galaxy name",
    "isophote profile PDF",
    "GB visual class",
    "GB visual notes",
    "PE classification",
    "PE profile class",
    "PE profile label",
    "VPD classification",
    "VPD profile class",
    "VPD profile label",
    "sra_classification",
    "sra_classification_detail",
    "left_shoulder_found",
    "right_shoulder_found",
    "failed_extrema",
    "d_extrema",
    "d2_extrema",
    "roc_minima",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LINK_FONT = Font(color="0563C1", underline="single")


def default_workbook_path(pc_name: str) -> Path:
    return PC_RESEARCH_FOLDERS[pc_name] / "Shoulder_Recognition_Erwin" / WORKBOOK_NAME


def default_isophote_dir(pc_name: str) -> Path:
    return PC_RESEARCH_FOLDERS[pc_name] / "Erwin" / "isophote_output" / "individual"


def reorder_sheet_columns(sheet, preferred_columns: list[str]) -> None:
    headers = [sheet.cell(row=1, column=index).value for index in range(1, sheet.max_column + 1)]
    header_to_index = {str(header): index for index, header in enumerate(headers, start=1) if header}

    ordered_headers = list(preferred_columns)
    ordered_headers.extend(header for header in headers if header and header not in ordered_headers)

    rows = []
    for row_index in range(2, sheet.max_row + 1):
        row = {}
        for header, column_index in header_to_index.items():
            row[header] = sheet.cell(row=row_index, column=column_index).value
        rows.append(row)

    if sheet.max_column:
        sheet.delete_cols(1, sheet.max_column)

    for column_index, header in enumerate(ordered_headers, start=1):
        sheet.cell(row=1, column=column_index).value = header
        for row_index, row in enumerate(rows, start=2):
            sheet.cell(row=row_index, column=column_index).value = row.get(header)


def column_index_by_header(sheet) -> dict[str, int]:
    return {
        str(sheet.cell(row=1, column=column_index).value).strip(): column_index
        for column_index in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=column_index).value
    }


def add_visual_review_columns(sheet, isophote_dir: Path) -> None:
    headers = column_index_by_header(sheet)
    galaxy_column = headers["galaxy name"]
    link_column = headers["isophote profile PDF"]
    visual_class_column = headers["GB visual class"]
    visual_notes_column = headers["GB visual notes"]

    class_column_letter = sheet.cell(row=1, column=visual_class_column).column_letter
    validation = DataValidation(
        type="list",
        formula1='"Peak+Sh,Exp,Flat-top (FT),Two-slope (2S),Unclear"',
        allow_blank=True,
    )
    validation.error = "Choose one of the profile classes in the dropdown."
    validation.errorTitle = "Invalid profile class"
    validation.prompt = "Select your visual classification."
    validation.promptTitle = "GB visual class"
    sheet.data_validations.dataValidation = []
    sheet.add_data_validation(validation)
    validation.add(f"{class_column_letter}2:{class_column_letter}{sheet.max_row}")

    for row_index in range(2, sheet.max_row + 1):
        galaxy = str(sheet.cell(row=row_index, column=galaxy_column).value or "").strip()
        link_cell = sheet.cell(row=row_index, column=link_column)
        link_cell.value = ""
        link_cell.hyperlink = None

        if galaxy:
            profile_path = isophote_dir / f"{galaxy}_isophote_axes.pdf"
            if profile_path.exists():
                link_cell.value = "open profile PDF"
                link_cell.hyperlink = profile_path.as_uri()
                link_cell.font = LINK_FONT

        sheet.cell(row=row_index, column=visual_notes_column).alignment = Alignment(wrap_text=True, vertical="top")


def style_headers(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    for column_cells in sheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        width = min(max(max_length + 2, 12), 42)
        sheet.column_dimensions[column_cells[0].column_letter].width = width


def format_workbook(workbook_path: Path, isophote_dir: Path) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook["Classifications"]
    reorder_sheet_columns(sheet, PREFERRED_COLUMNS)
    add_visual_review_columns(sheet, isophote_dir)
    style_headers(sheet)

    headers = column_index_by_header(sheet)
    if "isophote profile PDF" in headers:
        sheet.column_dimensions[sheet.cell(row=1, column=headers["isophote profile PDF"]).column_letter].width = 20
    if "GB visual class" in headers:
        sheet.column_dimensions[sheet.cell(row=1, column=headers["GB visual class"]).column_letter].width = 18
    if "GB visual notes" in headers:
        sheet.column_dimensions[sheet.cell(row=1, column=headers["GB visual notes"]).column_letter].width = 34

    if "Definitions" in workbook.sheetnames:
        style_headers(workbook["Definitions"])

    backup_path = workbook_path.with_name(
        f"{workbook_path.stem}.backup_{datetime.now():%Y%m%d_%H%M%S}{workbook_path.suffix}"
    )
    shutil.copy2(workbook_path, backup_path)
    workbook.save(workbook_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Reorder and style the PE/VPD/SRA classification workbook."
    )
    parser.add_argument("--pc", choices=sorted(PC_RESEARCH_FOLDERS), default="Laptop")
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--isophote-dir", type=Path, default=None)
    args = parser.parse_args()
    if args.workbook is None:
        args.workbook = default_workbook_path(args.pc)
    if args.isophote_dir is None:
        args.isophote_dir = default_isophote_dir(args.pc)
    return args


def main() -> int:
    args = parse_args()
    if not args.workbook.exists():
        raise FileNotFoundError(f"Could not find workbook: {args.workbook}")
    if not args.isophote_dir.exists():
        raise FileNotFoundError(f"Could not find isophote directory: {args.isophote_dir}")
    format_workbook(args.workbook, args.isophote_dir)
    print(f"Formatted workbook: {args.workbook}")
    print(f"Linked isophote PDFs from: {args.isophote_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
