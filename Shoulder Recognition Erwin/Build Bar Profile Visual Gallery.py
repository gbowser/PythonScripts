#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build a local HTML gallery for visually classifying bar-profile PDFs."""

from __future__ import annotations

import argparse
import html
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
from pathlib import Path
from urllib.parse import quote, unquote

from openpyxl import load_workbook


PC_RESEARCH_FOLDERS = {
    "Laptop": Path(r"C:\Users\gordo\Dropbox\Public Documents\UCLAN\MSc Research"),
    "Desktop": Path(r"D:\Dropbox\Public Documents\UCLAN\MSc Research"),
}

WORKBOOK_NAME = "PE_VPD_galaxy_classifications_with_definitions.xlsx"
CLASS_OPTIONS = ["", "Peak+Sh", "Exp", "Flat-top (FT)", "Two-slope (2S)", "Unclear"]


def default_workbook_path(pc_name: str) -> Path:
    return PC_RESEARCH_FOLDERS[pc_name] / "Shoulder_Recognition_Erwin" / WORKBOOK_NAME


def default_isophote_dir(pc_name: str) -> Path:
    return PC_RESEARCH_FOLDERS[pc_name] / "Erwin" / "isophote_output" / "individual"


def default_output_path(pc_name: str) -> Path:
    return PC_RESEARCH_FOLDERS[pc_name] / "Shoulder_Recognition_Erwin" / "bar_profile_visual_gallery.html"


def row_value(row: dict[str, object], key: str) -> str:
    return str(row.get(key) or "").strip()


def read_workbook_rows(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook["Classifications"]
    headers = [sheet.cell(row=1, column=column).value for column in range(1, sheet.max_column + 1)]
    rows = []
    for row_index in range(2, sheet.max_row + 1):
        row = {
            str(header).strip(): sheet.cell(row=row_index, column=column).value
            for column, header in enumerate(headers, start=1)
            if header
        }
        if row_value(row, "galaxy name"):
            rows.append(row)
    return rows


def build_gallery_data(rows: list[dict[str, object]], isophote_dir: Path, serve_pdfs: bool = False) -> list[dict[str, str]]:
    gallery_rows = []
    for row in rows:
        galaxy = row_value(row, "galaxy name")
        profile_path = isophote_dir / f"{galaxy}_isophote_axes.pdf"
        if serve_pdfs and profile_path.exists():
            pdf_url = f"/pdf/{quote(profile_path.name)}"
        else:
            pdf_url = profile_path.as_uri() if profile_path.exists() else ""
        gallery_rows.append(
            {
                "galaxy": galaxy,
                "pdf": pdf_url,
                "pe": row_value(row, "PE profile label"),
                "vpd": row_value(row, "VPD profile label"),
                "sra": row_value(row, "sra_classification"),
                "visualClass": row_value(row, "GB visual class"),
                "notes": row_value(row, "GB visual notes"),
            }
        )
    return gallery_rows


def column_index_by_header(sheet) -> dict[str, int]:
    return {
        str(sheet.cell(row=1, column=column_index).value).strip(): column_index
        for column_index in range(1, sheet.max_column + 1)
        if sheet.cell(row=1, column=column_index).value
    }


def update_workbook_visual_review(workbook_path: Path, galaxy: str, visual_class: str, notes: str) -> None:
    workbook = load_workbook(workbook_path)
    sheet = workbook["Classifications"]
    headers = column_index_by_header(sheet)
    galaxy_column = headers["galaxy name"]
    visual_class_column = headers["GB visual class"]
    notes_column = headers["GB visual notes"]

    for row_index in range(2, sheet.max_row + 1):
        row_galaxy = str(sheet.cell(row=row_index, column=galaxy_column).value or "").strip()
        if row_galaxy == galaxy:
            sheet.cell(row=row_index, column=visual_class_column).value = visual_class
            sheet.cell(row=row_index, column=notes_column).value = notes
            workbook.save(workbook_path)
            return

    raise ValueError(f"Galaxy not found in workbook: {galaxy}")


def render_html(gallery_rows: list[dict[str, str]], output_path: Path) -> str:
    data_json = json.dumps(gallery_rows, ensure_ascii=True)
    options_json = json.dumps(CLASS_OPTIONS, ensure_ascii=True)
    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Bar Profile Visual Classification Gallery</title>
  <style>
    :root {{
      color-scheme: light;
      font-family: Arial, Helvetica, sans-serif;
      background: #f4f6f8;
      color: #172033;
    }}
    body {{
      margin: 0;
    }}
    header {{
      position: sticky;
      top: 0;
      z-index: 2;
      display: grid;
      grid-template-columns: minmax(220px, 1fr) auto auto;
      gap: 12px;
      align-items: center;
      padding: 14px 18px;
      background: #17365d;
      color: white;
      box-shadow: 0 2px 10px rgba(0, 0, 0, 0.16);
    }}
    h1 {{
      margin: 0;
      font-size: 20px;
      font-weight: 700;
    }}
    input, textarea, button {{
      font: inherit;
    }}
    #search {{
      width: min(360px, 100%);
      padding: 8px 10px;
      border: 1px solid #9fb4cf;
      border-radius: 4px;
    }}
    button {{
      padding: 8px 12px;
      border: 1px solid #0f2746;
      border-radius: 4px;
      background: #ffffff;
      color: #17365d;
      cursor: pointer;
      font-weight: 700;
    }}
    .toggle {{
      display: inline-flex;
      align-items: center;
      gap: 7px;
      white-space: nowrap;
      font-weight: 700;
    }}
    #save-status {{
      min-width: 118px;
      font-size: 13px;
      color: #dcecff;
    }}
    #version-label {{
      font-size: 12px;
      color: #b9cce5;
      white-space: nowrap;
    }}
    main {{
      display: grid;
      grid-template-columns: repeat(auto-fit, minmax(520px, 1fr));
      gap: 16px;
      padding: 16px;
    }}
    article {{
      display: grid;
      grid-template-rows: auto auto auto;
      gap: 10px;
      background: white;
      border: 1px solid #d7dee8;
      border-radius: 6px;
      padding: 12px;
    }}
    .meta {{
      display: grid;
      grid-template-columns: 1fr auto;
      gap: 8px;
      align-items: start;
    }}
    .meta-actions {{
      display: grid;
      gap: 8px;
      justify-items: end;
    }}
    .reveal-one {{
      padding: 6px 9px;
      border-color: #9fb4cf;
      font-size: 13px;
      font-weight: 700;
      white-space: nowrap;
    }}
    .reveal-one[disabled] {{
      cursor: default;
      opacity: 0.68;
    }}
    h2 {{
      margin: 0;
      font-size: 18px;
    }}
    .tags {{
      display: flex;
      flex-wrap: wrap;
      gap: 6px;
      margin-top: 8px;
    }}
    .tag {{
      padding: 3px 7px;
      border-radius: 4px;
      background: #e7eef7;
      color: #17365d;
      font-size: 13px;
    }}
    .tag.hidden-classifier {{
      background: #f0f2f5;
      color: #697386;
    }}
    iframe {{
      width: 100%;
      aspect-ratio: 2.45 / 1;
      height: auto;
      min-height: 260px;
      max-height: 520px;
      border: 1px solid #c8d2df;
      border-radius: 4px;
      background: #f8fafc;
    }}
    .missing {{
      display: grid;
      place-items: center;
      border: 1px dashed #aeb9c8;
      border-radius: 4px;
      color: #6b7280;
    }}
    .review {{
      display: grid;
      grid-template-columns: minmax(160px, 220px) 1fr;
      gap: 10px;
    }}
    .review textarea {{
      width: 100%;
      box-sizing: border-box;
      border: 1px solid #b6c2d1;
      border-radius: 4px;
      padding: 8px;
      color: #111827;
      background: #ffffff;
    }}
    .review textarea {{
      min-height: 42px;
      resize: vertical;
    }}
    .class-picker {{
      position: relative;
    }}
    .class-picker-button {{
      width: 100%;
      min-height: 42px;
      box-sizing: border-box;
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 8px;
      border: 1px solid #b6c2d1;
      border-radius: 4px;
      padding: 8px 10px;
      background: #ffffff;
      color: #111827;
      font-weight: 400;
      text-align: left;
    }}
    .class-picker-button::after {{
      content: "v";
      color: #111827;
      font-size: 13px;
    }}
    .class-picker-menu {{
      position: absolute;
      left: 0;
      right: 0;
      bottom: calc(100% + 4px);
      z-index: 5;
      display: none;
      overflow: hidden;
      border: 1px solid #b6c2d1;
      border-radius: 4px;
      background: #ffffff;
      box-shadow: 0 8px 20px rgba(15, 39, 70, 0.18);
    }}
    .class-picker.open .class-picker-menu {{
      display: block;
    }}
    .class-option {{
      display: block;
      width: 100%;
      border: 0;
      border-radius: 0;
      padding: 9px 10px;
      background: #ffffff;
      color: #111827;
      font-weight: 400;
      text-align: left;
    }}
    .class-option:hover, .class-option:focus {{
      background: #e7eef7;
      outline: none;
    }}
    .class-option.selected {{
      background: #17365d;
      color: #ffffff;
    }}
    a {{
      color: #0b57a3;
      font-weight: 700;
    }}
    @media (max-width: 720px) {{
      header, main, article, .review, .meta {{
        grid-template-columns: 1fr;
      }}
      .meta-actions {{
        justify-items: start;
      }}
      main {{
        padding: 10px;
      }}
    }}
  </style>
</head>
<body>
  <header>
    <h1>Bar Profile Visual Classification Gallery</h1>
    <input id="search" type="search" placeholder="Search galaxy, PE, VPD, SRA">
    <label class="toggle"><input id="reveal-classifiers" type="checkbox"> Reveal Classifiers</label>
    <button id="export" type="button">Export CSV</button>
    <span id="save-status" aria-live="polite"></span>
    <span id="version-label">v2026-06-25 hide-all</span>
  </header>
  <main id="gallery"></main>
  <script>
    const rows = {data_json};
    const classOptions = {options_json};
    const storagePrefix = "barProfileVisual:";
    const gallery = document.getElementById("gallery");
    const search = document.getElementById("search");
    const revealClassifiers = document.getElementById("reveal-classifiers");
    const saveStatus = document.getElementById("save-status");
    const workbookSaveEnabled = window.location.protocol.startsWith("http");
    const saveTimers = new Map();

    function escapeHtml(value) {{
      return String(value ?? "").replace(/[&<>"']/g, char => ({{
        "&": "&amp;",
        "<": "&lt;",
        ">": "&gt;",
        '"': "&quot;",
        "'": "&#39;"
      }}[char]));
    }}

    function savedFor(galaxy) {{
      const saved = localStorage.getItem(storagePrefix + galaxy);
      return saved ? JSON.parse(saved) : null;
    }}

    function save(galaxy, visualClass, notes) {{
      localStorage.setItem(storagePrefix + galaxy, JSON.stringify({{ visualClass, notes }}));
    }}

    async function saveToWorkbook(galaxy, visualClass, notes) {{
      if (!workbookSaveEnabled) {{
        return;
      }}
      saveStatus.textContent = "Saving...";
      try {{
        const response = await fetch("/api/classification", {{
          method: "POST",
          headers: {{ "Content-Type": "application/json" }},
          body: JSON.stringify({{ galaxy, visualClass, notes }})
        }});
        if (!response.ok) {{
          throw new Error(await response.text());
        }}
        saveStatus.textContent = "Saved to workbook";
      }} catch (error) {{
        saveStatus.textContent = "Workbook save failed";
        console.error(error);
      }}
    }}

    function queueSaveToWorkbook(galaxy, visualClass, notes) {{
      if (!workbookSaveEnabled) {{
        return;
      }}
      window.clearTimeout(saveTimers.get(galaxy));
      saveTimers.set(galaxy, window.setTimeout(() => saveToWorkbook(galaxy, visualClass, notes), 500));
    }}

    function combinedText(row) {{
      return [row.galaxy, row.pe, row.vpd, row.sra].join(" ").toLowerCase();
    }}

    function embeddedPdfUrl(pdfUrl) {{
      return `${{pdfUrl}}#page=1&view=FitH&pagemode=none&toolbar=0`;
    }}

    function classLabel(value) {{
      return value || "Select...";
    }}

    function classOptionButtons(selectedValue) {{
      return classOptions.map(option => {{
        const selectedClass = option === selectedValue ? " selected" : "";
        return `<button type="button" class="class-option${{selectedClass}}" data-value="${{escapeHtml(option)}}">${{escapeHtml(classLabel(option))}}</button>`;
      }}).join("");
    }}

    function updatePickerSelection(picker, pickerButton, selectedValue) {{
      picker.dataset.value = selectedValue;
      pickerButton.textContent = classLabel(selectedValue);
      for (const optionButton of picker.querySelectorAll(".class-option")) {{
        optionButton.classList.toggle("selected", optionButton.dataset.value === selectedValue);
      }}
      picker.classList.remove("open");
    }}

    function classifierTags(row, showClassifiers) {{
      const peTag = showClassifiers
        ? `<span class="tag">PE: ${{escapeHtml(row.pe || "-")}}</span>`
        : `<span class="tag hidden-classifier">PE hidden</span>`;
      const vpdTag = showClassifiers
        ? `<span class="tag">VPD: ${{escapeHtml(row.vpd || "-")}}</span>`
        : `<span class="tag hidden-classifier">VPD hidden</span>`;
      const sraTag = showClassifiers
        ? `<span class="tag">SRA: ${{escapeHtml(row.sra || "-")}}</span>`
        : `<span class="tag hidden-classifier">SRA hidden</span>`;
      return `${{peTag}}${{vpdTag}}${{sraTag}}`;
    }}

    function classifiersVisibleFor(article) {{
      return revealClassifiers.checked || article.dataset.classifiersRevealed === "true";
    }}

    function updateClassifierCard(article) {{
      const row = rows.find(item => item.galaxy === article.dataset.galaxy);
      if (!row) {{
        return;
      }}
      const showClassifiers = classifiersVisibleFor(article);
      const tags = article.querySelector(".tags");
      if (tags) {{
        tags.innerHTML = classifierTags(row, showClassifiers);
      }}
      const revealButton = article.querySelector(".reveal-one");
      if (revealButton) {{
        revealButton.textContent = showClassifiers ? "Hide Classifiers" : "Reveal Classifiers";
        revealButton.disabled = revealClassifiers.checked;
        if (revealClassifiers.checked) {{
          revealButton.textContent = "Global Reveal On";
        }}
      }}
    }}

    function updateClassifierCards() {{
      for (const article of gallery.querySelectorAll("article")) {{
        updateClassifierCard(article);
      }}
    }}

    function updateSearchVisibility() {{
      const term = search.value.trim().toLowerCase();
      for (const article of gallery.querySelectorAll("article")) {{
        const row = rows.find(item => item.galaxy === article.dataset.galaxy);
        article.hidden = Boolean(term && (!row || !combinedText(row).includes(term)));
      }}
    }}

    function render() {{
      gallery.innerHTML = rows.map(row => {{
        const saved = savedFor(row.galaxy);
        const visualClass = saved?.visualClass ?? row.visualClass ?? "";
        const notes = saved?.notes ?? row.notes ?? "";
        const pdf = row.pdf
          ? `<iframe src="${{escapeHtml(embeddedPdfUrl(row.pdf))}}"></iframe>`
          : `<div class="missing">No matching isophote PDF found</div>`;
        const openLink = row.pdf ? `<a href="${{escapeHtml(row.pdf)}}" target="_blank">Open PDF</a>` : "";
        const showClassifiers = false;
        return `<article data-galaxy="${{escapeHtml(row.galaxy)}}">
          <div class="meta">
            <div>
              <h2>${{escapeHtml(row.galaxy)}}</h2>
              <div class="tags">
                ${{classifierTags(row, showClassifiers)}}
              </div>
            </div>
            <div class="meta-actions">
              <div>${{openLink}}</div>
              <button type="button" class="reveal-one">Reveal Classifiers</button>
            </div>
          </div>
          ${{pdf}}
          <div class="review">
            <div class="class-picker" data-field="class" data-value="${{escapeHtml(visualClass)}}">
              <button type="button" class="class-picker-button">${{escapeHtml(classLabel(visualClass))}}</button>
              <div class="class-picker-menu">${{classOptionButtons(visualClass)}}</div>
            </div>
            <textarea data-field="notes" placeholder="Notes"></textarea>
          </div>
        </article>`;
      }}).join("");

      for (const article of gallery.querySelectorAll("article")) {{
        const galaxy = article.dataset.galaxy;
        const row = rows.find(item => item.galaxy === galaxy);
        const saved = savedFor(galaxy);
        const picker = article.querySelector('.class-picker[data-field="class"]');
        const pickerButton = picker.querySelector(".class-picker-button");
        const textarea = article.querySelector('textarea[data-field="notes"]');
        const revealOneButton = article.querySelector(".reveal-one");
        picker.dataset.value = saved?.visualClass ?? row.visualClass ?? "";
        textarea.value = saved?.notes ?? row.notes ?? "";
        revealOneButton.addEventListener("click", () => {{
          article.dataset.classifiersRevealed = article.dataset.classifiersRevealed === "true" ? "false" : "true";
          updateClassifierCard(article);
        }});
        pickerButton.addEventListener("click", () => {{
          for (const otherPicker of gallery.querySelectorAll(".class-picker.open")) {{
            if (otherPicker !== picker) {{
              otherPicker.classList.remove("open");
            }}
          }}
          picker.classList.toggle("open");
        }});
        for (const optionButton of picker.querySelectorAll(".class-option")) {{
          optionButton.addEventListener("click", () => {{
            updatePickerSelection(picker, pickerButton, optionButton.dataset.value);
            save(galaxy, picker.dataset.value, textarea.value);
            saveToWorkbook(galaxy, picker.dataset.value, textarea.value);
          }});
        }}
        textarea.addEventListener("focus", () => {{
          picker.classList.remove("open");
        }});
        textarea.addEventListener("input", () => {{
          save(galaxy, picker.dataset.value, textarea.value);
          queueSaveToWorkbook(galaxy, picker.dataset.value, textarea.value);
        }});
      }}
      updateSearchVisibility();
      updateClassifierCards();
    }}

    function csvEscape(value) {{
      return `"${{String(value ?? "").replace(/"/g, '""')}}"`;
    }}

    document.getElementById("export").addEventListener("click", () => {{
      const lines = [["galaxy", "GB visual class", "GB visual notes"].map(csvEscape).join(",")];
      for (const row of rows) {{
        const saved = savedFor(row.galaxy);
        lines.push([row.galaxy, saved?.visualClass ?? row.visualClass ?? "", saved?.notes ?? row.notes ?? ""].map(csvEscape).join(","));
      }}
      const blob = new Blob([lines.join("\\n")], {{ type: "text/csv;charset=utf-8" }});
      const url = URL.createObjectURL(blob);
      const link = document.createElement("a");
      link.href = url;
      link.download = "gb_visual_bar_profile_classifications.csv";
      link.click();
      URL.revokeObjectURL(url);
    }});

    search.addEventListener("input", updateSearchVisibility);
    revealClassifiers.addEventListener("change", updateClassifierCards);
    document.addEventListener("click", event => {{
      if (!event.target.closest(".class-picker")) {{
        for (const picker of gallery.querySelectorAll(".class-picker.open")) {{
          picker.classList.remove("open");
        }}
      }}
    }});
    render();
  </script>
</body>
</html>
"""


def serve_gallery(workbook_path: Path, isophote_dir: Path, host: str, port: int) -> None:
    class GalleryHandler(BaseHTTPRequestHandler):
        def log_message(self, format: str, *args: object) -> None:
            return

        def send_text(self, status: int, text: str, content_type: str = "text/plain; charset=utf-8") -> None:
            body = text.encode("utf-8")
            self.send_response(status)
            self.send_header("Content-Type", content_type)
            self.send_header("Cache-Control", "no-store, max-age=0")
            self.send_header("Pragma", "no-cache")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)

        def do_GET(self) -> None:
            if self.path in {"/", "/bar_profile_visual_gallery.html"}:
                rows = read_workbook_rows(workbook_path)
                gallery_rows = build_gallery_data(rows, isophote_dir, serve_pdfs=True)
                self.send_text(200, render_html(gallery_rows, workbook_path), "text/html; charset=utf-8")
                return

            if self.path.startswith("/pdf/"):
                pdf_name = unquote(self.path.removeprefix("/pdf/").split("?", 1)[0].split("#", 1)[0])
                pdf_path = (isophote_dir / pdf_name).resolve()
                try:
                    pdf_path.relative_to(isophote_dir.resolve())
                except ValueError:
                    self.send_text(404, "PDF not found")
                    return
                if not pdf_path.exists():
                    self.send_text(404, "PDF not found")
                    return
                body = pdf_path.read_bytes()
                self.send_response(200)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Cache-Control", "no-store, max-age=0")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)
                return

            self.send_text(404, "Not found")

        def do_POST(self) -> None:
            if self.path != "/api/classification":
                self.send_text(404, "Not found")
                return

            try:
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                galaxy = str(payload.get("galaxy") or "").strip()
                visual_class = str(payload.get("visualClass") or "").strip()
                notes = str(payload.get("notes") or "").strip()
                if not galaxy:
                    raise ValueError("Missing galaxy name")
                if visual_class and visual_class not in CLASS_OPTIONS:
                    raise ValueError(f"Invalid visual class: {visual_class}")
                update_workbook_visual_review(workbook_path, galaxy, visual_class, notes)
                self.send_text(200, json.dumps({"ok": True}), "application/json; charset=utf-8")
            except Exception as exc:
                self.send_text(500, str(exc))

    server = ThreadingHTTPServer((host, port), GalleryHandler)
    print(f"Serving gallery at http://{host}:{port}/")
    print(f"Workbook saves go to: {workbook_path}")
    print("Press Ctrl+C to stop the gallery server.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopped gallery server.")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build an HTML gallery for bar-profile visual classification.")
    parser.add_argument("--pc", choices=sorted(PC_RESEARCH_FOLDERS), default="Laptop")
    parser.add_argument("--workbook", type=Path, default=None)
    parser.add_argument("--isophote-dir", type=Path, default=None)
    parser.add_argument("--output", type=Path, default=None)
    parser.add_argument("--serve", action="store_true", help="Serve the gallery locally and write visual classifications back to the workbook.")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=8765)
    args = parser.parse_args()
    if args.workbook is None:
        args.workbook = default_workbook_path(args.pc)
    if args.isophote_dir is None:
        args.isophote_dir = default_isophote_dir(args.pc)
    if args.output is None:
        args.output = default_output_path(args.pc)
    return args


def main() -> int:
    args = parse_args()
    if not args.workbook.exists():
        raise FileNotFoundError(f"Could not find workbook: {args.workbook}")
    if not args.isophote_dir.exists():
        raise FileNotFoundError(f"Could not find isophote directory: {args.isophote_dir}")

    rows = read_workbook_rows(args.workbook)
    if args.serve:
        serve_gallery(args.workbook, args.isophote_dir, args.host, args.port)
        return 0

    gallery_rows = build_gallery_data(rows, args.isophote_dir)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(render_html(gallery_rows, args.output), encoding="utf-8")

    linked_count = sum(1 for row in gallery_rows if row["pdf"])
    print(f"Wrote gallery: {args.output}")
    print(f"Rows: {len(gallery_rows)}; linked PDFs: {linked_count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
